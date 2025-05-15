import io
import csv
import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from .statistics import PresenceStatisticsService

class ExportService:
    """Service pour l'exportation des données de présence"""
    
    @staticmethod
    def export_presence_count_by_date(start_date=None, end_date=None, classe_id=None, format='xlsx'):
        """
        Exporte le nombre de présences par jour dans une période donnée
        """
        # Récupérer les données
        data = PresenceStatisticsService.get_presence_count_by_date(
            start_date=start_date,
            end_date=end_date,
            classe_id=classe_id
        )
        
        # Titre du rapport
        title = "Rapport de présences par jour"
        if start_date and end_date:
            title += f" du {start_date} au {end_date}"
        elif start_date:
            title += f" depuis le {start_date}"
        elif end_date:
            title += f" jusqu'au {end_date}"
            
        # Colonnes du rapport
        headers = ['Date', 'Nombre de présences']
        
        # Données formatées
        rows = [[item['day'], item['count']] for item in data]
        
        # Exporter selon le format demandé
        if format == 'xlsx':
            return ExportService._export_to_excel(title, headers, rows)
        elif format == 'csv':
            return ExportService._export_to_csv(title, headers, rows)
        elif format == 'pdf':
            return ExportService._export_to_pdf(title, headers, rows)
        else:
            raise ValueError(f"Format d'exportation non supporté: {format}")
    
    @staticmethod
    def export_presence_count_by_class(start_date=None, end_date=None, format='xlsx'):
        """
        Exporte le nombre de présences par classe dans une période donnée
        """
        # Récupérer les données
        data = PresenceStatisticsService.get_presence_count_by_class(
            start_date=start_date,
            end_date=end_date
        )
        
        # Titre du rapport
        title = "Rapport de présences par classe"
        if start_date and end_date:
            title += f" du {start_date} au {end_date}"
        elif start_date:
            title += f" depuis le {start_date}"
        elif end_date:
            title += f" jusqu'au {end_date}"
            
        # Colonnes du rapport
        headers = ['Classe', 'Nombre de présences']
        
        # Données formatées
        rows = [[item['classe_nom'], item['count']] for item in data]
        
        # Exporter selon le format demandé
        if format == 'xlsx':
            return ExportService._export_to_excel(title, headers, rows)
        elif format == 'csv':
            return ExportService._export_to_csv(title, headers, rows)
        elif format == 'pdf':
            return ExportService._export_to_pdf(title, headers, rows)
        else:
            raise ValueError(f"Format d'exportation non supporté: {format}")
    
    @staticmethod
    def export_attendance_rate_by_student(start_date=None, end_date=None, classe_id=None, format='xlsx'):
        """
        Exporte le taux de présence par étudiant dans une période donnée
        """
        # Récupérer les données
        data = PresenceStatisticsService.get_attendance_rate_by_student(
            start_date=start_date,
            end_date=end_date,
            classe_id=classe_id
        )
        
        # Titre du rapport
        title = "Rapport de taux de présence par étudiant"
        if start_date and end_date:
            title += f" du {start_date} au {end_date}"
        elif start_date:
            title += f" depuis le {start_date}"
        elif end_date:
            title += f" jusqu'au {end_date}"
            
        # Colonnes du rapport
        headers = ['Nom', 'Prénom', 'Classe', 'Présences', 'Jours ouvrables', 'Taux de présence (%)']
        
        # Données formatées
        rows = [
            [
                item['etudiant_nom'],
                item['etudiant_prenom'],
                item['classe_nom'] or 'Non assigné',
                item['presence_count'],
                item['working_days'],
                f"{item['attendance_rate']}%"
            ]
            for item in data
        ]
        
        # Exporter selon le format demandé
        if format == 'xlsx':
            return ExportService._export_to_excel(title, headers, rows)
        elif format == 'csv':
            return ExportService._export_to_csv(title, headers, rows)
        elif format == 'pdf':
            return ExportService._export_to_pdf(title, headers, rows)
        else:
            raise ValueError(f"Format d'exportation non supporté: {format}")
    
    @staticmethod
    def export_absence_alerts(threshold=70, days=30, format='xlsx'):
        """
        Exporte les alertes d'absence
        """
        # Récupérer les données
        data = PresenceStatisticsService.get_absence_alerts(
            threshold=threshold,
            days=days
        )
        
        # Titre du rapport
        title = f"Alertes d'absentéisme (seuil: {threshold}%, période: {days} jours)"
            
        # Colonnes du rapport
        headers = ['Nom', 'Prénom', 'Classe', 'Présences', 'Jours ouvrables', 'Taux de présence (%)']
        
        # Données formatées
        rows = [
            [
                item['etudiant_nom'],
                item['etudiant_prenom'],
                item['classe_nom'] or 'Non assigné',
                item['presence_count'],
                item['working_days'],
                f"{item['attendance_rate']}%"
            ]
            for item in data
        ]
        
        # Exporter selon le format demandé
        if format == 'xlsx':
            return ExportService._export_to_excel(title, headers, rows)
        elif format == 'csv':
            return ExportService._export_to_csv(title, headers, rows)
        elif format == 'pdf':
            return ExportService._export_to_pdf(title, headers, rows)
        else:
            raise ValueError(f"Format d'exportation non supporté: {format}")
    
    @staticmethod
    def _export_to_excel(title, headers, rows):
        """
        Exporte les données au format Excel
        """
        # Créer un workbook et une feuille
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport"
        
        # Ajouter le titre
        ws.merge_cells('A1:F1')
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        # Ajouter les en-têtes
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        
        # Ajouter les données
        for row_idx, row_data in enumerate(rows, 4):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Ajuster la largeur des colonnes
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}_{datetime.date.today()}.xlsx"'
        
        # Sauvegarder le workbook dans la réponse
        wb.save(response)
        
        return response
    
    @staticmethod
    def _export_to_csv(title, headers, rows):
        """
        Exporte les données au format CSV
        """
        # Créer la réponse HTTP
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}_{datetime.date.today()}.csv"'
        
        # Créer le writer CSV
        writer = csv.writer(response)
        
        # Écrire le titre
        writer.writerow([title])
        writer.writerow([])  # Ligne vide
        
        # Écrire les en-têtes
        writer.writerow(headers)
        
        # Écrire les données
        for row in rows:
            writer.writerow(row)
        
        return response
    
    @staticmethod
    def _export_to_pdf(title, headers, rows):
        """
        Exporte les données au format PDF
        """
        # Créer un buffer pour le PDF
        buffer = io.BytesIO()
        
        # Créer le document PDF
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        # Styles pour le document
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        # Éléments du document
        elements = []
        
        # Ajouter le titre
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))
        
        # Préparer les données pour le tableau
        table_data = [headers] + rows
        
        # Créer le tableau
        table = Table(table_data)
        
        # Style du tableau
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        # Appliquer le style au tableau
        table.setStyle(table_style)
        
        # Ajouter le tableau au document
        elements.append(table)
        
        # Construire le document
        doc.build(elements)
        
        # Créer la réponse HTTP
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}_{datetime.date.today()}.pdf"'
        
        return response
